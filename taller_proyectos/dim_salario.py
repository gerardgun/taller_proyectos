from openpyxl import load_workbook
import re

ARRAY_KEYS = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 15]


def check_information(key_words, cadena):
    count = 0
    for key in key_words:
        if key in cadena:
            count = count + 1

    return count

def get_prom(list_amounts):
    return sum(list_amounts)/list_amounts.__len__()


def check_dollar(cadena):
    try:
        if '$' in str(cadena):
            return True
    except Exception:
        pass
    return False


def check_soles(cadena):
    if 'SOLES' in str(cadena) or 'S/.' in str(cadena):
        return True
    return False

def check_hourly_rate(cadena):
    if 'HR' in cadena:
        return True
    return False

def check_month_payments(cadena):
    if 'MENSUAL' in cadena:
        return True
    return False


def get_monto_format(cadena):
    number_list = []
    for numb in re.findall(r'\d+,\d+|\d+\.\d+,\d\d|\d+', cadena):
        number_format = numb.replace('.', '').replace(',00', '').replace('.00', '').replace(',', '')
        number_list.append(float(number_format))
    return number_list


def clean_data(cadena):
    array_to_ignore = [',', 'DE', 'Y', 'EN', 'A', 'EL', '(', ')', 'QUE', 'CON', 'LAS', 'LOS', 'UN', 'POR', 'LO', 'LA',
                       'OTRAS', 'SER', 'PARA', 'O', '/', 'SUS', 'AL', 'U', 'UNA', 'ASI', '&', '']

    for ignore in array_to_ignore:
        cadena.replace(ignore, ' ')

    return cadena


def update_excel():
    wb = load_workbook(filename='/home/gerard/Documentos/last_excel2.xlsx')
    sheet = wb.get_sheet_by_name('Hoja3')
    for row in range(2, sheet.max_row + 1):
        cell_id = sheet.cell(row=row, column=1).value
        cell_description = sheet.cell(row=row, column=2).value
        cell_FK = sheet.cell(row=row, column=3)
        cell_query_fk = sheet.cell(row=row, column=11)
        data_result = []
        if cell_id not in ARRAY_KEYS and cell_FK.value is None:
            #check ocurrences
            if check_dollar(cell_description):
                if check_hourly_rate(cell_description):
                    list_amount = get_monto_format(cell_description)
                    amount = get_prom(list_amount)
                    if 0 < amount < 10:
                        sheet.cell(row=row, column=3).value = 1
                        sheet.cell(row=row, column=11).value = 1
                    if 11 <= amount <= 20:
                        sheet.cell(row=row, column=3).value = 2
                        sheet.cell(row=row, column=11).value = 2
                    if 21 <= amount <= 30:
                        sheet.cell(row=row, column=3).value = 3
                        sheet.cell(row=row, column=11).value = 3
                    if amount >= 31:
                        sheet.cell(row=row, column=3).value = 4
                        sheet.cell(row=row, column=11).value = 4
                else:
                    sheet.cell(row=row, column=3).value = 10
                    sheet.cell(row=row, column=11).value = 10
            if check_soles(cell_description):
                if check_month_payments(cell_description):
                    list_amount = get_monto_format(cell_description)
                    amount = get_prom(list_amount)
                    if 0 < amount <= 1000:
                        sheet.cell(row=row, column=3).value = 5
                        sheet.cell(row=row, column=11).value = 5
                    if 1001 <= amount <= 2000:
                        sheet.cell(row=row, column=3).value = 6
                        sheet.cell(row=row, column=11).value = 6
                    if 2001 <= amount < 3500:
                        sheet.cell(row=row, column=3).value = 7
                        sheet.cell(row=row, column=11).value = 7
                    if 3501 <= amount <= 5000:
                        sheet.cell(row=row, column=3).value = 8
                        sheet.cell(row=row, column=11).value = 8
                    if amount >= 5001:
                        sheet.cell(row=row, column=3).value = 9
                        sheet.cell(row=row, column=11).value = 9
                else:
                    sheet.cell(row=row, column=3).value = 11
                    sheet.cell(row=row, column=11).value = 11


    wb.save('/home/gerard/Documentos/last_excel2.xlsx')