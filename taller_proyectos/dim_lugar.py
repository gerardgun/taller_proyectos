from openpyxl import load_workbook
LIMA = ['LIMA', 'PERU'] # 1
SAN_ISIDRO = ['SAN ISIDRO'] # 9
MIRAFLORES = ['MIRAFLORES'] # 13
CERCADO = ['CERCADO DE LIMA']  # 14
SANTIAGO_SURCO = ['SANTIAGO', 'SURCO'] # 15
CALLAO = ['CALLAO'] # 21
INDIA = ['INDIA'] #23
JESUS = ['JESUS MARIA'] #24
UK = ['UNITED KINGDOM', 'UK'] #25
SAN_BORJA = ['SAN BORJA']  #28
AREQUIPA = ['AREQUIPA'] #30
TRUJILLO = ['TRUJILLO', 'LA LIBERTAD'] #33
SANTIAGO = ['SANTIAGO', 'CHILE'] #38
AUSTRALIA = ['AUSTRALIA'] #37
VICTO_LARCO = ['VICTOR LARCO', 'TRUJILLO'] #41
USA = ['USA', 'UNITED STATES']  #42
MADGALENA = ['MADGALENA'] #44
BARRANCO = ['BARRANCO'] #46
NIGERIA = ['NIGERIA'] #47
ATE = ['ATE'] #49
SPAIN = ['ESPAÑA', 'SPAIN'] #51
SAN_JUAN = ['SAN JUAN DE LURIGANCHO'] #56
SINGAPORE = ['SINGAPORE'] #62
TUMBES = ['TUMBES'] #65
JAEN = ['CAJAMARCA', 'JAEN'] #67
CHORRILLOS = ['CHORRILLOS'] #69
VILLA = ['VILLA EL SALVADOR'] #70
MEXICO = ['MEXICO', 'CIUDAD'] #74
COLOMBIA = ['COLOMBIA', 'BOGOTA'] #76
MALAYSIA =['MALAYSIA'] #77
SANTA_ANITA =['SANTA ANITA'] #78
CHICLAYO = ['CHICLAYO', 'LAMBAYEQUE'] #79
ARABE_EMIRATES = ['UNITED ARAB EMIRATES'] #82
LINCE = ['LINCE'] #84
CANADA = ['CANADA'] #86
ARGENTINA =['BUENOS AIRES', 'ARGENTINA'] #88
TUMBES =  ['TUMBES', 'PERU'] #90
ECUADOR = ['QUITO', 'ECUADOR'] #92
RUSSIA = ['RUSSIAN FEDERATION'] #96
SAN_MIGUEL =['SAN MIGUEL'] #97
LA_MOLINA = ['LA MOLINA'] #98
FRANCIA = ['FRANCE'] #115
GERMANY = ['GERMANY'] #117
TURKIA = ['TURKEY'] #119
ARABIA = ['SAUDI ARABIA'] #121
LINCE = ['LINCE'] #122
VENEZUELA = ['VENEZUELA'] #152
EGYPT = ['EGYPT'] #159
LIBIA = ['LEBANON'] #161
CUSCO = ['CUSCO', 'CUZCO'] #167
SOUT_AFRICA = ['SOUTH AFRICA'] #190
CHINA = ['CHINA'] #206
SWITZERLAND = ['SWITZERLAND'] #215
BRAZIL =['BRAZIL'] #245
ITALIA = ['ITALY'] #246

DATA = [
    {
    'id': 1,
    'info': LIMA
    },
    {
        'id': 9,
        'info': SAN_ISIDRO
    },
    {
        'id': 14,
        'info': CERCADO
    },
    {
        'id': 15,
        'info': SANTIAGO_SURCO
    },
    {
        'id': 21,
        'info': CALLAO
    },
    {
        'id': 23,
        'info': INDIA
    },
{
        'id': 25,
        'info': UK
    },
{
        'id': 28,
        'info': SAN_BORJA
    },
{
        'id': 30,
        'info': AREQUIPA
    },
{
        'id': 33,
        'info': TRUJILLO
    },
{
        'id': 38,
        'info': SANTIAGO
    },
{
        'id': 37,
        'info': AUSTRALIA
    },
{
        'id': 41,
        'info': VICTO_LARCO
    },
{
        'id': 42,
        'info': USA
    },
{
        'id': 44,
        'info': MADGALENA
    },
{
        'id': 46,
        'info': BARRANCO
    },
{
        'id': 47,
        'info': NIGERIA
    },
{
        'id': 49,
        'info': ATE
    },
{
        'id': 51,
        'info': SPAIN
    },
{
        'id': 56,
        'info': SAN_JUAN
    },
{
        'id': 62,
        'info': SINGAPORE
    },
{
        'id': 67,
        'info': JAEN
    },
{
        'id': 67,
        'info': JAEN
    },
{
        'id': 69,
        'info': CHORRILLOS
    },
{
        'id': 70,
        'info': VILLA
    },
{
        'id': 74,
        'info': MEXICO
    },
{
        'id': 76,
        'info': COLOMBIA
    },
{
        'id': 77,
        'info': MALAYSIA
    },
{
        'id': 78,
        'info': SANTA_ANITA
    },
{
        'id': 79,
        'info': CHICLAYO
    },
{
        'id': 82,
        'info': ARABE_EMIRATES
    },
{
        'id': 84,
        'info': LINCE
    },
{
        'id': 86,
        'info': CANADA
    },
{
        'id': 88,
        'info': ARGENTINA
    },{
        'id': 90,
        'info': TUMBES
    },
{
        'id': 92,
        'info': ECUADOR
    },
{
        'id': 96,
        'info': RUSSIA
    },
{
        'id': 97,
        'info': SAN_MIGUEL
    },
{
        'id': 98,
        'info': LA_MOLINA
    },
{
        'id': 115,
        'info': FRANCIA
    },
{
        'id': 117,
        'info': GERMANY
    },
{
        'id': 119,
        'info': TURKIA
    },
{
        'id': 121,
        'info': ARABIA
    },
{
        'id': 122,
        'info': LINCE
    },
{
        'id': 152,
        'info': VENEZUELA
    },
{
        'id': 159,
        'info': EGYPT
    },
{
        'id': 161,
        'info': LIBIA
    },
{
        'id': 167,
        'info': CUSCO
    },
{
        'id': 190,
        'info': SOUT_AFRICA
    },
{
        'id': 206,
        'info': CHINA
    },
{
        'id': 215,
        'info': SWITZERLAND
    },
{
        'id': 245,
        'info': BRAZIL
    },
{
        'id': 246,
        'info': ITALIA
    },

]


ARRAY_KEYS = [1, 9, 13, 14, 15, 21, 23, 24, 25, 28, 30, 33, 37, 38, 41, 42, 44, 46, 47,
              49, 51, 56, 62, 65, 67, 69, 70, 74, 76, 77, 78, 79, 82, 84, 86, 88, 90, 92, 96,
              97, 98, 115, 117, 119, 121, 122, 152, 159, 161, 167, 190, 206, 215, 245, 246]


def check_information(key_words, cadena):
    count = 0
    for key in key_words:
        if not cadena:
            return 0
        if key in cadena:
            count = count + 1

    return count



def update_excel():
    wb = load_workbook(filename='/home/gerard/Documentos/last_excel3_2.xlsx')
    sheet = wb.get_sheet_by_name('Hoja2')
    for row in range(2, sheet.max_row + 1):
        cell_id = sheet.cell(row=row, column=1).value
        cell_description = sheet.cell(row=row, column=2).value
        cell_FK = sheet.cell(row=row, column=3)
        cell_query_fk = sheet.cell(row=row, column=11)
        data_result = []
        if cell_id not in ARRAY_KEYS and cell_FK.value is None:
            #check ocurrences
            for data in DATA:
                count = check_information(key_words=data['info'], cadena=cell_description)
                data_result.append({'id': data['id'], 'count': count})
            # count = check_information(key_words=TECNICO_BACHILLER, cadena=cell_description)
            # data_result.append({'id': 1, 'count': count})
            # count = check_information(key_words=TECNICO_TITULADO, cadena=cell_description)
            # data_result.append({'id': 2, 'count': count})
            # count = check_information(key_words=BACHILLER_UNIVERSITARIO_TITULO, cadena=cell_description)
            # data_result.append({'id': 3, 'count': count})
            # count = check_information(key_words=PROFESIONAL_TECNICO, cadena=cell_description)
            # data_result.append({'id': 7, 'count': count})

            # count = check_information(key_words=INGENIERO, cadena=cell_description)
            # data_result.append({'id': 8, 'count': count})
            # count = check_information(key_words=EGRESADO_BACHILLER, cadena=cell_description)
            # data_result.append({'id': 11, 'count': count})


        if data_result.__len__() > 0:
            new_list = sorted(data_result, key=lambda k: k['count'])
            if new_list[-1]['count'] > 0:
                print(sheet.cell(row, column=1).value, sheet.cell(row, column=2).value, new_list[-1]['id'])
                sheet.cell(row=row, column=3).value = new_list[-1]['id']
                sheet.cell(row=row, column=11).value = new_list[-1]['id']

    wb.save('/home/gerard/Documentos/last_excel3_2.xlsx')