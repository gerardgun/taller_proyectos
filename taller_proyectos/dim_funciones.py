from openpyxl import load_workbook

TEST_UNITARIOS = ['PRUEBAS', 'TEST', 'AUTOMATION', 'UNIT', 'AUTOMATIZA', 'UNITARIAS']
DESARROLLO_IMPLEMETACION = ['DESARROLLO', 'EQUIPO', 'IMPLEMENTAR', 'IMPLEMENTACION', 'DESARROLLAR', 'WEB', 'PROGRAMAR',
                            'APLICACIONES', 'FRONT', 'BACK', 'ESPECIFICACIONES', 'SISTEMAS', 'SOFTWARE', 'HERRAMIENTAS','SYSTEM',
                            'PROGRAMAS', 'MANEJO', 'INTEGRATION', 'DEVELOPMENT', 'APP', 'DEVELOP', 'MIGRATE', 'MOBILE', 'IOS', 'ANDROID']
ESTIMACIONES = ['ESTIMAR', 'ESTIMACIONES', 'TIEMPO', 'ENTREGA', 'CALCULAR', 'EVALUAR']
MANTENIMIENTO = ['MANTENIMIENTO', 'ERRORES', 'FIX', 'OPTIMIZACION', 'CORRECCIONES', 'FUNCIONALIDADES', 'ACTUALIZAR',
                 'NUEVAS', 'BUG', 'ERROR', 'ENHANCEMENTS', ' TECHNICAL', 'ISSUE', 'OPTIMIZE', 'FOUND']
INVESTIGACION = ['INVESTIGACION', 'INVESTIGAR', 'NUEVAS', 'PROPONER', 'TECNOLOGIAS', 'TENDENCIAS']
GESTION = ['LIDERAR', 'GESTIONAR', 'PROYECTOS', 'PLANIFICAR', 'PLANIFICACION', 'COORDINAR', 'PROYECTOS', 'GERENCIAL',
           'PRESENTACION']
DEVOPS = ['DESPLIEGUES', ' UAT ', 'PRODUCCION', 'PRODUCTION', 'CLOUD', 'AZURE', 'AWS', 'CONFIGURACION', 'SERVIDORES',
           'AUTOMATIZACION', 'PIPELINES', 'DESARROLLAR', 'DEVOPS', 'SYSTEMS']
DOCUMENTADOR = ['DOCUMENTOS', 'DOCUMENTAR', 'ELABORACION', 'MANUALES', 'USUARIO', 'ELABORAR', 'FINAL',]
REQUERIMIENTOS = ['RECOGER', 'LEVANTAR', 'REQUERIMIENTOS', 'REUNION', 'REUNIRSE', 'CLIENTES']
QA = [' QA ', 'QUALITY', 'PROBAR', 'PRUEBAS', 'CALIDAD', 'ERRORES', 'INCONSISTENCIAS', 'ENCONTRAR', 'ENCONTRADAS', 'CODIFICACIONES']
SEGURIDAD = ['USUARIO', 'CIBERSEGURIDAD', 'ACCESOS', 'PRIVILEGIOS', 'BACKUPS', 'AUTENTICACION', 'INFORMACION', 'RESPALDO']
CAPACITACION = ['CLIENTES', 'CAPACITACION', 'USUARIOS', 'FINALES', 'REUNIRSE', 'ATENCION', 'REUNIONES', 'ASESORAR',
                'MANEJO', 'USO']
SOPORTE = ['FORMATEO', 'INSTALACION', 'COMPUTO', 'EQUIPOS', 'SOPORTE', 'TECNICO', 'MANTENIMIENTO', 'LAPTOP',
           'HARDWARE', ' PC ', 'WINDOWS']
MOCKUPS = ['MAQUETAR', 'MOCKUPS', 'DISEÑO', 'PRESENTACION', 'UX', ' UI ', 'UI/UX', 'DISEÑAR', 'PROTOTIPO', 'INTERFAZ', 'USUARIO']
ANALISTA = ['ANALIZAR', 'MODELAR', 'DIAGRAMAR', 'REQUERIMIENTOS', 'PROCESOS', 'EVALUAR', 'MODELOS', 'CREACION', 'ANALITICOS']
BIG_DATA = ['BIG','DATA', 'MACHINE', 'LEARNING', ' ML ', ' AI ', ' IA ', 'INTELIGENCIA', 'ARTIFICIAL']
PM = ['LIDERAR', 'MOTIVAR', 'AVANCES', 'REVISAR', 'COORDINAR']
GENERALES = ['REALIZAR', 'ASIGNADAS', 'FUNCIONES', 'RESPONSABILIDADES', 'INMEDIATO', 'COMUNICARSE', 'EXPRESAR', 'COMUNICAR',
             'NEGOCIO', 'COLABORAR', 'EQUIPO' ,'COMUNICARTE', 'PROGRAMACION', 'PYTHON', 'JAVA', 'MYSQL', 'POSTGRES', 'C#',
             'PHP', 'VUE', 'REACT NATIVE', 'FLUTTER', 'SWIFT', 'MANEJO', 'MOODLE', 'PLATAFORMA', 'HERRAMIENTAS',
             ]
ARRAY_KEYS = [1, 2, 4, 5, 6, 7, 8, 9, 10, 11, 14, 42, 53, 103, 137, 260, 382, 455]


def check_information(key_words, cadena):
    count = 0
    for key in key_words:
        if key in cadena:
            count = count + 1

    return count



def clean_data(cadena):
    array_to_ignore = [',', 'DE', 'Y', 'EN', 'A', 'EL', '(', ')', 'QUE', 'CON', 'LAS', 'LOS', 'UN', 'POR', 'LO', 'LA',
                       'OTRAS', 'SER', 'PARA', 'O', '/', 'SUS', 'AL', 'U', 'UNA', 'ASI', '&', '']

    for ignore in array_to_ignore:
        cadena.replace(ignore, ' ')

    return cadena


def update_excel():
    wb = load_workbook(filename='/home/gerard/Escritorio/presentacion2.xlsx')
    sheet = wb.get_sheet_by_name('funcion')
    for row in range(2, sheet.max_row + 1):
        cell_id = sheet.cell(row=row, column=1).value
        cell_description = sheet.cell(row=row, column=2).value
        cell_FK = sheet.cell(row=row, column=3)
        cell_query_fk = sheet.cell(row=row, column=11)
        data_result = []
        if cell_id not in ARRAY_KEYS and cell_FK.value is None:
            #check ocurrences
            count = check_information(key_words=TEST_UNITARIOS, cadena=cell_description)
            data_result.append({'id': 1, 'count': count})
            count = check_information(key_words=DESARROLLO_IMPLEMETACION, cadena=cell_description)
            data_result.append({'id': 2, 'count': count})
            count = check_information(key_words=ESTIMACIONES, cadena=cell_description)
            data_result.append({'id': 4, 'count': count})
            count = check_information(key_words=MANTENIMIENTO, cadena=cell_description)
            data_result.append({'id': 5, 'count': count})
            count = check_information(key_words=INVESTIGACION, cadena=cell_description)
            data_result.append({'id': 6, 'count': count})
            count = check_information(key_words=GESTION, cadena=cell_description)
            data_result.append({'id': 7, 'count': count})
            count = check_information(key_words=DEVOPS, cadena=cell_description)
            data_result.append({'id': 8, 'count': count})
            count = check_information(key_words=GENERALES, cadena=cell_description)
            data_result.append({'id': 9, 'count': count})
            count = check_information(key_words=DOCUMENTADOR, cadena=cell_description)
            data_result.append({'id': 10, 'count': count})
            count = check_information(key_words=REQUERIMIENTOS, cadena=cell_description)
            data_result.append({'id': 11, 'count': count})
            count = check_information(key_words=QA, cadena=cell_description)
            data_result.append({'id': 14, 'count': count})
            count = check_information(key_words=SEGURIDAD, cadena=cell_description)
            data_result.append({'id': 42, 'count': count})
            count = check_information(key_words=CAPACITACION, cadena=cell_description)
            data_result.append({'id': 53, 'count': count})
            count = check_information(key_words=SOPORTE, cadena=cell_description)
            data_result.append({'id': 103, 'count': count})
            count = check_information(key_words=MOCKUPS, cadena=cell_description)
            data_result.append({'id': 137, 'count': count})
            count = check_information(key_words=ANALISTA, cadena=cell_description)
            data_result.append({'id': 260, 'count': count})
            count = check_information(key_words=BIG_DATA, cadena=cell_description)
            data_result.append({'id': 382, 'count': count})
            count = check_information(key_words=PM, cadena=cell_description)
            data_result.append({'id': 455, 'count': count})

        if data_result.__len__() > 0:
            new_list = sorted(data_result, key=lambda k: k['count'])
            if new_list[-1]['count'] > 0:
                print(sheet.cell(row, column=1).value, sheet.cell(row, column=2).value, new_list[-1]['id'])
                sheet.cell(row=row, column=3).value = new_list[-1]['id']
                sheet.cell(row=row, column=11).value = new_list[-1]['id']

    wb.save('/home/gerard/Escritorio/presentacion2.xlsx')
